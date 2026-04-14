from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
ws = wb.active
ws.title = "AI 점수 기준표"

header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='2F5496')
title_font = Font(name='Arial', bold=True, size=13, color='2F5496')
sub_font = Font(name='Arial', italic=True, size=9, color='666666')
data_font = Font(name='Arial', size=10)
score_high = PatternFill('solid', fgColor='C6EFCE')
score_mid = PatternFill('solid', fgColor='FFEB9C')
score_low = PatternFill('solid', fgColor='FFC7CE')
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
center = Alignment(horizontal='center', vertical='center')

def sfill(v):
    if v >= 80: return score_high
    elif v >= 50: return score_mid
    return score_low

def add_header(row, titles):
    for c, t in enumerate(titles, 1):
        cell = ws.cell(row=row, column=c, value=t)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

def set_cell(row, col, val, is_score=False):
    cell = ws.cell(row=row, column=col, value=val)
    cell.font = data_font
    cell.alignment = center
    cell.border = thin_border
    if is_score and isinstance(val, (int, float)):
        cell.fill = sfill(val)

def add_section(start_row, title, sub, rows_data):
    ws.merge_cells(start_column=1, end_column=6, start_row=start_row, end_row=start_row)
    ws.cell(row=start_row, column=1, value=title).font = title_font
    ws.cell(row=start_row, column=1).alignment = Alignment(horizontal='left')

    add_header(start_row + 1, ['등락율', '점수', '유동주식 대비 거래량(평소대비)', '점수', '거래량 변동(전봉대비)', '점수'])

    for i, (a, b, c, d, e, f) in enumerate(rows_data):
        r = start_row + 2 + i
        set_cell(r, 1, a)
        set_cell(r, 2, b, True) if isinstance(b, int) else set_cell(r, 2, b)
        set_cell(r, 3, c)
        set_cell(r, 4, d, True) if isinstance(d, int) else set_cell(r, 4, d)
        set_cell(r, 5, e)
        set_cell(r, 6, f, True) if isinstance(f, int) else set_cell(r, 6, f)

    end_row = start_row + 2 + len(rows_data)
    ws.cell(row=end_row, column=1, value=sub).font = sub_font
    return end_row + 2

# ── 일봉 - 전일대비 ──
rows_day = [
    ('+2~5%', 85, '5배 이상', 95, '300% 이상', 50),
    ('+1~2%', 65, '3~5배', 90, '200~300%', 35),
    ('+5~7%', 70, '2~3배', 80, '120~200%', 20),
    ('+7~9%', 30, '1.5~2배', 70, '120% 미만', 0),
    ('+9% 초과', 10, '1~1.5배', 55, '', ''),
    ('0~1%', 50, '0.5~1배', 35, '', ''),
    ('하락', 30, '0.5배 미만', 20, '', ''),
]
next_row = add_section(1, "일봉 - 전일대비 (DayChangeCondition)", "* DayChangeCondition(W:15) / TurnoverCondition / VolumeSurgeCondition", rows_day)

# ── 일봉 - 시가대비 ──
rows_open = [
    ('+2~5%', 85, '5배 이상', 95, '300% 이상', 50),
    ('+1~2%', 65, '3~5배', 90, '200~300%', 35),
    ('+5~7%', 60, '2~3배', 80, '120~200%', 20),
    ('+7~9%', 30, '1.5~2배', 70, '120% 미만', 0),
    ('+9% 초과', 10, '1~1.5배', 55, '', ''),
    ('0~1%', 50, '0.5~1배', 35, '', ''),
    ('하락', 30, '0.5배 미만', 20, '', ''),
]
next_row = add_section(next_row, "일봉 - 시가대비 (OpenChangeCondition)", "* OpenChangeCondition(W:20) / TurnoverCondition / VolumeSurgeCondition", rows_open)

# ── 60분봉 ──
rows_60 = [
    ('+3~5%', 90, '5배 이상', 95, '300% 이상', 50),
    ('+1~3%', 70, '3~5배', 90, '200~300%', 35),
    ('+0.5~1%', 55, '2~3배', 80, '120~200%', 20),
    ('0~0.5%', 40, '1.5~2배', 70, '120% 미만', 0),
    ('+5% 초과', 40, '1~1.5배', 55, '', ''),
    ('-1~0%', 10, '0.5~1배', 35, '', ''),
    ('-1% 미만', 0, '0.5배 미만', 20, '', ''),
]
next_row = add_section(next_row, "60분봉 (MinuteChangeCondition)", "* MinuteChangeCondition(W:50) / MinuteTurnoverCondition(W:50) / VolumeSurgeCondition", rows_60)

# ── 15분봉 ──
next_row = add_section(next_row, "15분봉 (Min15ChangeCondition) - 타이밍 확인용", "* Min15ChangeCondition(W:25) / Min15TurnoverCondition(W:25) / 60분봉보다 가중치 낮게 설정", rows_60)

# ── 가중치 요약 ──
ws.merge_cells(start_column=1, end_column=6, start_row=next_row, end_row=next_row)
ws.cell(row=next_row, column=1, value="고려사항 가중치(W) 요약").font = title_font

add_header(next_row + 1, ['조건명', '봉기준', 'W', '비고', '', ''])
weights = [
    ('60분봉 등락율', '60분', 50, '+3~5% 스윗스팟'),
    ('60분봉 유동주식 거래비중', '60분', 50, '높을수록 가점'),
    ('15분봉 등락율', '15분', 25, '타이밍 확인'),
    ('15분봉 유동주식 거래비중', '15분', 25, '타이밍 확인'),
    ('전일대비 등락', '일봉', 15, '+7% 이상 감점'),
    ('시가대비 등락', '일봉', 20, '+7% 이상 감점'),
    ('체결강도', '-', 25, '100 이상'),
    ('수급 2일 이상', '-', 20, '외인/기관'),
    ('외국인 순매수', '-', 20, ''),
    ('기관 순매수', '-', 20, ''),
    ('큰손 수급', '-', 15, '종합 세력'),
    ('프로그램 매수', '-', 10, ''),
    ('거래대금', '-', 15, ''),
    ('RSI', '-', 15, '70 이하'),
    ('볼린저', '-', 10, '과열 체크'),
    ('20일선 지지', '-', 15, ''),
    ('60일선 지지', '-', 15, ''),
    ('박스권 돌파', '-', 15, ''),
    ('상승테마', '-', 20, ''),
    ('업종지수', '-', 15, ''),
    ('코스피코스닥', '-', 15, ''),
    ('뉴스', '-', 15, ''),
    ('과거 유사패턴', '-', 15, 'AI 학습'),
]
for i, (name, tf, w, note) in enumerate(weights):
    r = next_row + 2 + i
    set_cell(r, 1, name)
    set_cell(r, 2, tf)
    set_cell(r, 3, w)
    set_cell(r, 4, note)

ws.column_dimensions['A'].width = 16
ws.column_dimensions['B'].width = 8
ws.column_dimensions['C'].width = 34
ws.column_dimensions['D'].width = 8
ws.column_dimensions['E'].width = 24
ws.column_dimensions['F'].width = 8

output = r'C:\StockTrader\AI_점수기준표.xlsx'
wb.save(output)
print(f"저장 완료: {output}")
