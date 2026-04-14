from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "매수 고려사항 분석"

header_font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='2F5496')
cat_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
cat_fill = PatternFill('solid', fgColor='4472C4')
data_font = Font(name='Arial', size=10)
warn_fill = PatternFill('solid', fgColor='FFF2CC')
dup_fill = PatternFill('solid', fgColor='FCE4D6')
keep_fill = PatternFill('solid', fgColor='E2EFDA')
core_fill = PatternFill('solid', fgColor='C6EFCE')
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)

def hdr(row, titles):
    for c, t in enumerate(titles, 1):
        cell = ws.cell(row=row, column=c, value=t)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

def cat_row(row, title):
    ws.merge_cells(start_column=1, end_column=8, start_row=row, end_row=row)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = cat_font
    cell.fill = cat_fill
    cell.alignment = Alignment(horizontal='left', vertical='center')
    cell.border = thin_border
    for c in range(2, 9):
        ws.cell(row=row, column=c).border = thin_border
        ws.cell(row=row, column=c).fill = cat_fill

def add(row, num, name, w, why, pro, con, dup, rec, rec_fill=None):
    vals = [num, name, w, why, pro, con, dup, rec]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.font = data_font
        cell.border = thin_border
        if c <= 3:
            cell.alignment = center
        else:
            cell.alignment = left_wrap
    if rec_fill:
        ws.cell(row=row, column=8).fill = rec_fill

# 제목
ws.merge_cells('A1:H1')
ws.cell(row=1, column=1, value="매수 고려사항 전체 분석표").font = Font(name='Arial', bold=True, size=14, color='2F5496')
ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')

# 헤더
hdr(3, ['#', '조건명', 'W', '왜 넣었나', '장점', '단점', '중복/정리', '추천'])

r = 4

# 수급/세력
cat_row(r, "수급 / 세력"); r += 1
add(r, 1, "체결강도 100이상", 25, "실시간 매수세 vs 매도세 판단", "현재 매수 압력을 직접 확인", "순간 스파이크에 속을 수 있음", "단독", "유지", keep_fill); r += 1
add(r, 2, "수급 2일 이상", 20, "외인/기관 연속 매수 확인", "단기 추세 신뢰도 높음", "3,4번과 겹침 (종합 수급)", "3,4번과 중복", "제거 추천 (3,4번으로 대체)", dup_fill); r += 1
add(r, 3, "외국인 순매수", 20, "외국인 단독 매수 확인", "외국인 움직임 따로 판단", "2번에 이미 포함", "2번과 중복", "유지 (세분화)", keep_fill); r += 1
add(r, 4, "기관 순매수", 20, "기관 단독 매수 확인", "기관 움직임 따로 판단", "2번에 이미 포함", "2번과 중복", "유지 (세분화)", keep_fill); r += 1
add(r, 5, "큰손 수급", 15, "외인+기관+프로그램+기금 종합", "세력 전체 흐름 파악", "2,3,4번과 또 겹침", "2,3,4와 중복", "유지 (종합 관점)", keep_fill); r += 1
add(r, 6, "프로그램 매수", 10, "기관 알고리즘 매매 동향", "기관 자동매매 포착", "5번 큰손 수급에 포함됨", "5번과 중복", "W 낮게 유지 or 제거", warn_fill); r += 1

# 거래량
cat_row(r, "거래량"); r += 1
add(r, 7, "거래대금", 15, "거래대금 충분한 종목만 진입", "유동성 부족 종목 걸러냄", "소형주 불리", "단독", "유지", keep_fill); r += 1
add(r, 8, "60분봉 유동주식 거래비중", 50, "평소 대비 거래량 집중도", "세력 유입 포착 핵심", "W:50 높음", "단독", "핵심 유지 (절대 빼지 말것)", core_fill); r += 1

# 기술적 지표
cat_row(r, "기술적 지표"); r += 1
add(r, 9, "RSI 70이하", 15, "과매수 구간 진입 방지", "고점 추격매수 방어", "강세장에서 좋은 종목도 걸러짐", "단독", "유지", keep_fill); r += 1
add(r, 10, "볼린저", 10, "밴드 상단 과열 체크", "통계적 과열 판단", "RSI와 비슷한 역할", "RSI와 유사", "유지 가능 (관점 다름)", keep_fill); r += 1
add(r, 11, "20일선 지지", 15, "단기 추세선 위에 있는지", "추세 확인 기본", "ACF에서 이미 이평선 필터 중", "ACF와 중복 가능", "유지 (ACF는 고정, 여기는 점수)", keep_fill); r += 1
add(r, 12, "60일선 지지", 15, "중기 추세선 지지", "중기 추세 확인", "단타에는 불필요할 수 있음", "단독", "유지", keep_fill); r += 1
add(r, 13, "박스권 상단 돌파", 15, "20일 고점 돌파 = 신고가", "돌파 매매 핵심 시그널", "가짜 돌파에 속을 수 있음", "단독", "유지", keep_fill); r += 1

# 등락/가격
cat_row(r, "등락 / 가격"); r += 1
add(r, 14, "시가대비 등락", 20, "장중 흐름 확인", "당일 움직임 판단", "15번과 비슷한 역할", "15번과 유사", "둘 중 하나만 써도 됨", warn_fill); r += 1
add(r, 15, "전일대비 등락", 15, "전일 대비 등락폭", "하루 기준 움직임", "14번과 비슷한 역할", "14번과 유사", "둘 중 하나만 써도 됨", warn_fill); r += 1
add(r, 16, "60분봉 등락율", 50, "60분봉 +3~5% 스윗스팟", "핵심 조건", "W:50 높음", "단독", "핵심 유지 (절대 빼지 말것)", core_fill); r += 1
add(r, 17, "15분봉 등락율", 25, "타이밍 확인", "진입 시점 정밀화", "노이즈 가능", "단독", "유지", keep_fill); r += 1
add(r, 18, "15분봉 유동주식 거래비중", 25, "15분봉 거래량 확인", "실시간 수급 확인", "노이즈 가능", "단독", "유지", keep_fill); r += 1

# 테마/시장
cat_row(r, "테마 / 시장"); r += 1
add(r, 19, "상승테마 소속", 20, "테마주 확인", "시장 관심 종목 포착", "테마 데이터 정확도", "단독", "유지", keep_fill); r += 1
add(r, 20, "업종지수 상승", 15, "업종 전체 흐름", "업종 동반 상승 확인", "개별주와 무관할 수 있음", "단독", "유지", keep_fill); r += 1
add(r, 21, "코스피 코스닥", 15, "시장 전체 분위기", "하락장 매수 방어", "개별주와 무관할 수 있음", "단독", "유지", keep_fill); r += 1
add(r, 22, "뉴스 긍정", 15, "뉴스 감성 분석", "호재 포착", "뉴스 데이터 정확도", "단독", "유지", keep_fill); r += 1

# AI
cat_row(r, "AI 학습"); r += 1
add(r, 23, "과거 유사패턴", 15, "AI 학습 기반 예측", "데이터 쌓이면 강력", "초기에는 데이터 부족", "단독", "유지 (장기 투자)", keep_fill); r += 1

# 정리 요약
r += 1
ws.merge_cells(start_column=1, end_column=8, start_row=r, end_row=r)
ws.cell(row=r, column=1, value="정리 추천 요약").font = Font(name='Arial', bold=True, size=13, color='2F5496')
r += 1

hdr(r, ['구분', '', '내용', '', '', '', '', '']); r += 1
ws.merge_cells(start_column=1, end_column=2, start_row=r, end_row=r)
ws.merge_cells(start_column=3, end_column=8, start_row=r, end_row=r)
ws.cell(row=r, column=1, value="중복 심한 것").font = Font(name='Arial', bold=True, size=10)
ws.cell(row=r, column=1).fill = dup_fill
ws.cell(row=r, column=1).border = thin_border
ws.cell(row=r, column=3, value="2번(수급2일) + 3번(외국인) + 4번(기관) + 5번(큰손) + 6번(프로그램) → 5개 모두 수급. 2번 제거, 3,4,5 유지 추천").font = data_font
ws.cell(row=r, column=3).border = thin_border
ws.cell(row=r, column=3).alignment = left_wrap
r += 1

ws.merge_cells(start_column=1, end_column=2, start_row=r, end_row=r)
ws.merge_cells(start_column=3, end_column=8, start_row=r, end_row=r)
ws.cell(row=r, column=1, value="유사한 것").font = Font(name='Arial', bold=True, size=10)
ws.cell(row=r, column=1).fill = warn_fill
ws.cell(row=r, column=1).border = thin_border
ws.cell(row=r, column=3, value="14번(시가대비) + 15번(전일대비) → 비슷한 역할. 하나만 남겨도 됨").font = data_font
ws.cell(row=r, column=3).border = thin_border
ws.cell(row=r, column=3).alignment = left_wrap
r += 1

ws.merge_cells(start_column=1, end_column=2, start_row=r, end_row=r)
ws.merge_cells(start_column=3, end_column=8, start_row=r, end_row=r)
ws.cell(row=r, column=1, value="핵심 조건").font = Font(name='Arial', bold=True, size=10)
ws.cell(row=r, column=1).fill = core_fill
ws.cell(row=r, column=1).border = thin_border
ws.cell(row=r, column=3, value="8번(60분봉 거래비중 W:50), 16번(60분봉 등락율 W:50) → 절대 빼면 안됨").font = data_font
ws.cell(row=r, column=3).border = thin_border
ws.cell(row=r, column=3).alignment = left_wrap

# 범례
r += 2
ws.cell(row=r, column=1, value="범례:").font = Font(name='Arial', bold=True, size=10)
r += 1
ws.cell(row=r, column=1, value="").fill = core_fill; ws.cell(row=r, column=1).border = thin_border
ws.cell(row=r, column=2, value="핵심 유지").font = data_font
r += 1
ws.cell(row=r, column=1, value="").fill = keep_fill; ws.cell(row=r, column=1).border = thin_border
ws.cell(row=r, column=2, value="유지").font = data_font
r += 1
ws.cell(row=r, column=1, value="").fill = warn_fill; ws.cell(row=r, column=1).border = thin_border
ws.cell(row=r, column=2, value="검토 필요").font = data_font
r += 1
ws.cell(row=r, column=1, value="").fill = dup_fill; ws.cell(row=r, column=1).border = thin_border
ws.cell(row=r, column=2, value="제거 추천").font = data_font

# 열 너비
ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 26
ws.column_dimensions['C'].width = 6
ws.column_dimensions['D'].width = 28
ws.column_dimensions['E'].width = 24
ws.column_dimensions['F'].width = 24
ws.column_dimensions['G'].width = 18
ws.column_dimensions['H'].width = 26

# 행 높이
for row in range(4, r + 1):
    ws.row_dimensions[row].height = 30

output = r'C:\StockTrader\AI_고려사항_분석.xlsx'
wb.save(output)
print(f"저장 완료: {output}")
