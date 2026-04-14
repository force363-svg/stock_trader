"""
백테스트 엔진
- historical_ohlcv 데이터로 매수/매도 전략 시뮬레이션
- 일별 스캔 → 조건 충족 시 가상 매수 → 매도 조건 체크 → 결과 집계
- 엑셀 리포트 출력
"""
import os
import sys
from datetime import datetime, timedelta
from ..db.database import get_connection

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def _get_report_path():
    if getattr(sys, 'frozen', False):
        base = os.path.dirname(os.path.dirname(sys.executable))
    else:
        base = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    report_dir = os.path.join(base, "reports")
    os.makedirs(report_dir, exist_ok=True)
    return report_dir


class Backtester:
    """과거 데이터 기반 전략 시뮬레이션"""

    def __init__(self, initial_cash=10_000_000, buy_amount=1_000_000, max_stocks=5):
        self.initial_cash = initial_cash
        self.buy_amount = buy_amount
        self.max_stocks = max_stocks

        # 시뮬레이션 상태
        self.cash = initial_cash
        self.holdings = {}  # {code: {"name", "buy_price", "buy_date", "qty"}}
        self.trades = []    # 완료된 매매 기록
        self.daily_equity = []  # 일별 자산 추적

        # 분할 손절 단계
        self._loss_stages = {}  # {code: done_stage}

        self._callback = None
        self._stop = False

    def stop(self):
        self._stop = True

    def run(self, start_date: str, end_date: str, callback=None):
        """
        백테스트 실행
        start_date/end_date: "YYYYMMDD" 형식
        callback: 진행 상태 콜백 (message: str)
        """
        self._callback = callback
        self._stop = False
        self.cash = self.initial_cash
        self.holdings = {}
        self.trades = []
        self.daily_equity = []
        self._loss_stages = {}

        dates = self._get_trading_dates(start_date, end_date)
        if not dates:
            self._msg("거래일 데이터 없음")
            return self._result()

        total = len(dates)
        self._msg(f"백테스트 시작: {start_date}~{end_date} ({total}일)")

        for i, date in enumerate(dates):
            if self._stop:
                break

            # 해당일 전 종목 데이터 로드
            all_stocks = self._load_day_data(date)
            if not all_stocks:
                continue

            # 1) 보유종목 매도 체크
            self._check_sell(date, all_stocks)

            # 2) 매수 후보 스캔
            if len(self.holdings) < self.max_stocks:
                self._check_buy(date, all_stocks)

            # 3) 일별 자산 기록
            equity = self._calc_equity(all_stocks)
            self.daily_equity.append({
                "date": date,
                "cash": self.cash,
                "equity": equity,
                "holdings": len(self.holdings),
            })

            if callback and (i + 1) % 20 == 0:
                pnl = (equity - self.initial_cash) / self.initial_cash * 100
                self._msg(f"[{date}] {i+1}/{total}일 | 자산:{equity:,.0f} ({pnl:+.1f}%)")

        # 미청산 포지션 강제 청산 (마지막 날 종가)
        if dates:
            last_stocks = self._load_day_data(dates[-1])
            if last_stocks:
                for code in list(self.holdings.keys()):
                    if code in last_stocks:
                        self._do_sell(code, last_stocks[code]["close"], dates[-1], "백테스트 종료")

        result = self._result()
        self._msg(f"백테스트 완료 | 매매:{result['total_trades']}회 승률:{result['win_rate']:.1f}% "
                  f"수익:{result['total_return']:+.1f}%")
        return result

    def _check_buy(self, date, all_stocks):
        """매수 조건 체크 — 4대 패턴 간이 시뮬레이션"""
        candidates = []

        for code, stock in all_stocks.items():
            if code in self.holdings:
                continue

            history = stock.get("history", [])
            if len(history) < 20:
                continue

            score = self._simple_buy_score(history)
            if score >= 70:
                candidates.append((code, stock, score))

        # 점수 높은 순으로 매수
        candidates.sort(key=lambda x: x[2], reverse=True)

        for code, stock, score in candidates:
            if len(self.holdings) >= self.max_stocks:
                break
            if self.cash < self.buy_amount:
                break

            price = stock["close"]
            if price <= 0:
                continue

            qty = self.buy_amount // price
            if qty <= 0:
                continue

            cost = qty * price
            self.cash -= cost
            self.holdings[code] = {
                "name": stock.get("name", code),
                "buy_price": price,
                "buy_date": date,
                "qty": qty,
                "score": score,
            }

    def _check_sell(self, date, all_stocks):
        """매도 조건 체크"""
        loss_stages = [(-3.0, 1/3), (-5.0, 1/3), (-7.0, 1.0)]

        for code in list(self.holdings.keys()):
            if code not in all_stocks:
                continue

            hold = self.holdings[code]
            cur_price = all_stocks[code]["close"]
            buy_price = hold["buy_price"]

            if buy_price <= 0:
                continue

            pnl = (cur_price - buy_price) / buy_price * 100

            # 분할 손절
            done_loss = self._loss_stages.get(code, -1)
            for stage_i, (threshold, ratio) in enumerate(loss_stages):
                if stage_i <= done_loss:
                    continue
                if pnl <= threshold:
                    if ratio >= 1.0:
                        self._do_sell(code, cur_price, date, f"분할손절 {threshold}%")
                    else:
                        sell_qty = max(1, int(hold["qty"] * ratio))
                        self._do_partial_sell(code, cur_price, date, sell_qty, f"분할손절 {threshold}%")
                        self._loss_stages[code] = stage_i
                    break

            if code not in self.holdings:
                continue

            # 매도 점수 간이 계산
            history = all_stocks[code].get("history", [])
            sell_score = self._simple_sell_score(history, pnl)

            if sell_score >= 70:
                self._do_sell(code, cur_price, date, f"매도점수 {sell_score:.0f}")
            elif pnl >= 10:
                # 10% 이상 수익 + 하락 전환
                if len(history) >= 2 and history[0]["close"] < history[1]["close"]:
                    self._do_sell(code, cur_price, date, f"수익실현 {pnl:+.1f}%")

    def _do_sell(self, code, price, date, reason):
        """전량 매도"""
        if code not in self.holdings:
            return
        hold = self.holdings[code]
        qty = hold["qty"]
        self.cash += qty * price
        pnl = (price - hold["buy_price"]) / hold["buy_price"] * 100

        self.trades.append({
            "code": code,
            "name": hold["name"],
            "buy_date": hold["buy_date"],
            "sell_date": date,
            "buy_price": hold["buy_price"],
            "sell_price": price,
            "qty": qty,
            "pnl": round(pnl, 2),
            "result": "WIN" if pnl > 0 else "LOSS",
            "reason": reason,
            "score": hold.get("score", 0),
        })
        del self.holdings[code]
        if code in self._loss_stages:
            del self._loss_stages[code]

    def _do_partial_sell(self, code, price, date, sell_qty, reason):
        """분할 매도"""
        if code not in self.holdings:
            return
        hold = self.holdings[code]
        sell_qty = min(sell_qty, hold["qty"])
        self.cash += sell_qty * price
        pnl = (price - hold["buy_price"]) / hold["buy_price"] * 100

        self.trades.append({
            "code": code,
            "name": hold["name"],
            "buy_date": hold["buy_date"],
            "sell_date": date,
            "buy_price": hold["buy_price"],
            "sell_price": price,
            "qty": sell_qty,
            "pnl": round(pnl, 2),
            "result": "WIN" if pnl > 0 else "LOSS",
            "reason": reason,
            "score": hold.get("score", 0),
        })

        hold["qty"] -= sell_qty
        if hold["qty"] <= 0:
            del self.holdings[code]
            if code in self._loss_stages:
                del self._loss_stages[code]

    def _simple_buy_score(self, history):
        """
        간이 매수 점수 (과거 데이터만으로 패턴 판단)
        history: 최신순 일봉 리스트 [{close, open, high, low, volume}, ...]
        """
        if len(history) < 20:
            return 0

        score = 0
        closes = [d["close"] for d in history]
        volumes = [d["volume"] for d in history]

        # 1) 이평선 정배열 (EMA5 > EMA20)
        ema5 = sum(closes[:5]) / 5
        ema20 = sum(closes[:20]) / 20
        if ema5 > ema20:
            score += 15

        # 2) 50일선 상승 (데이터 있으면)
        if len(closes) >= 50:
            sma50 = sum(closes[:50]) / 50
            sma50_prev = sum(closes[1:51]) / 50
            if sma50 > sma50_prev:
                score += 10

        # 3) 거래량 증가
        if len(volumes) >= 6:
            avg_vol = sum(volumes[1:6]) / 5
            if avg_vol > 0 and volumes[0] / avg_vol >= 1.5:
                score += 15

        # 4) 당일 양봉
        if history[0]["close"] > history[0]["open"]:
            score += 5

        # 5) 전일대비 상승 +2~7%
        if len(closes) >= 2 and closes[1] > 0:
            chg = (closes[0] - closes[1]) / closes[1] * 100
            if 2 <= chg <= 7:
                score += 15
            elif chg > 7:
                score += 5  # 과열

        # 6) 이격도 체크 (MA20 대비)
        if ema20 > 0:
            gap = (closes[0] - ema20) / ema20 * 100
            if gap <= 5:
                score += 10
            elif gap <= 10:
                score += 5

        # 7) 윗꼬리 없음
        candle_range = history[0]["high"] - history[0]["low"]
        if candle_range > 0:
            position = (history[0]["close"] - history[0]["low"]) / candle_range
            if position >= 0.5:
                score += 10

        return score

    def _simple_sell_score(self, history, pnl):
        """간이 매도 점수"""
        if len(history) < 5:
            return 0

        score = 0
        closes = [d["close"] for d in history]
        volumes = [d["volume"] for d in history]

        # 1) 거래량 감소 + 주가 하락
        if len(volumes) >= 6:
            avg_vol = sum(volumes[1:6]) / 5
            if avg_vol > 0:
                vol_ratio = volumes[0] / avg_vol
                price_chg = (closes[0] - closes[1]) / closes[1] * 100 if closes[1] > 0 else 0
                if vol_ratio <= 0.7 and price_chg < -1:
                    score += 25

        # 2) 5일선 이탈
        ema5 = sum(closes[:5]) / 5
        if closes[0] < ema5:
            score += 15

        # 3) 윗꼬리 감지
        candle_range = history[0]["high"] - history[0]["low"]
        if candle_range > 0:
            position = (history[0]["close"] - history[0]["low"]) / candle_range
            if position < 0.3:
                score += 20
                if pnl >= 5:
                    score += 10

        # 4) 음봉 + 거래량
        if history[0]["close"] < history[0]["open"] and volumes[0] > 0:
            score += 10

        # 5) 수익 구간별 매도 압력
        if pnl >= 15:
            score += 15
        elif pnl >= 10:
            score += 10

        return score

    def _load_day_data(self, date):
        """해당일 전 종목 OHLCV + 최근 60일 히스토리"""
        conn = get_connection()
        try:
            rows = conn.execute("""
                SELECT code, open, high, low, close, volume
                FROM historical_ohlcv
                WHERE date=? AND close > 0 AND volume > 0
            """, (date,)).fetchall()

            if not rows:
                return {}

            result = {}
            for r in rows:
                code = r["code"]
                # 최근 60일 히스토리
                hist = conn.execute("""
                    SELECT open, high, low, close, volume
                    FROM historical_ohlcv
                    WHERE code=? AND date<=?
                    ORDER BY date DESC LIMIT 60
                """, (code, date)).fetchall()

                result[code] = {
                    "open": r["open"],
                    "high": r["high"],
                    "low": r["low"],
                    "close": r["close"],
                    "volume": r["volume"],
                    "name": code,
                    "history": [dict(h) for h in hist],
                }
            return result
        finally:
            conn.close()

    def _get_trading_dates(self, start_date, end_date):
        """DB에서 실제 거래일 목록 조회"""
        conn = get_connection()
        try:
            rows = conn.execute("""
                SELECT DISTINCT date FROM historical_ohlcv
                WHERE date >= ? AND date <= ?
                ORDER BY date ASC
            """, (start_date, end_date)).fetchall()
            return [r["date"] for r in rows]
        finally:
            conn.close()

    def _calc_equity(self, all_stocks):
        """현재 총 자산 계산"""
        equity = self.cash
        for code, hold in self.holdings.items():
            if code in all_stocks:
                equity += hold["qty"] * all_stocks[code]["close"]
            else:
                equity += hold["qty"] * hold["buy_price"]
        return equity

    def _result(self):
        """결과 요약"""
        if not self.trades:
            return {
                "total_trades": 0, "wins": 0, "losses": 0, "win_rate": 0,
                "avg_pnl": 0, "max_profit": 0, "max_loss": 0,
                "total_return": 0, "trades": [], "daily_equity": self.daily_equity,
            }

        wins = [t for t in self.trades if t["pnl"] > 0]
        losses = [t for t in self.trades if t["pnl"] <= 0]
        pnls = [t["pnl"] for t in self.trades]

        final_equity = self.daily_equity[-1]["equity"] if self.daily_equity else self.initial_cash
        total_return = (final_equity - self.initial_cash) / self.initial_cash * 100

        return {
            "total_trades": len(self.trades),
            "wins": len(wins),
            "losses": len(losses),
            "win_rate": len(wins) / len(self.trades) * 100,
            "avg_pnl": sum(pnls) / len(pnls),
            "max_profit": max(pnls),
            "max_loss": min(pnls),
            "total_return": round(total_return, 2),
            "trades": self.trades,
            "daily_equity": self.daily_equity,
        }

    def _msg(self, text):
        print(f"[백테스트] {text}")
        if self._callback:
            self._callback(text)

    def export_excel(self, result=None):
        """백테스트 결과 엑셀 출력"""
        if not HAS_OPENPYXL:
            print("[백테스트] openpyxl 미설치")
            return None

        if result is None:
            result = self._result()

        wb = Workbook()
        header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="2D3436")
        data_font = Font(name="Arial", size=10)
        win_fill = PatternFill("solid", fgColor="DFF9FB")
        loss_fill = PatternFill("solid", fgColor="FFE0E0")
        border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )
        center = Alignment(horizontal="center", vertical="center")

        # ── Sheet 1: 요약 ──
        ws = wb.active
        ws.title = "백테스트 요약"
        ws.merge_cells("A1:D1")
        ws["A1"].value = f"백테스트 결과 ({datetime.now().strftime('%Y-%m-%d')})"
        ws["A1"].font = Font(name="Arial", bold=True, size=14)

        items = [
            ("총 매매", result["total_trades"], "회"),
            ("승리", result["wins"], "회"),
            ("패배", result["losses"], "회"),
            ("승률", f"{result['win_rate']:.1f}", "%"),
            ("평균 수익률", f"{result['avg_pnl']:+.2f}", "%"),
            ("최대 수익", f"{result['max_profit']:+.2f}", "%"),
            ("최대 손실", f"{result['max_loss']:+.2f}", "%"),
            ("총 수익률", f"{result['total_return']:+.2f}", "%"),
            ("초기자금", f"{self.initial_cash:,}", "원"),
        ]
        if result["daily_equity"]:
            final = result["daily_equity"][-1]["equity"]
            items.append(("최종자산", f"{final:,.0f}", "원"))

        for i, (label, value, unit) in enumerate(items, start=3):
            ws[f"A{i}"].value = label
            ws[f"A{i}"].font = Font(name="Arial", bold=True, size=10)
            ws[f"B{i}"].value = value
            ws[f"B{i}"].font = data_font
            ws[f"C{i}"].value = unit

        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 15

        # ── Sheet 2: 매매 내역 ──
        ws2 = wb.create_sheet("매매 내역")
        headers = ["종목코드", "매수일", "매도일", "매수가", "매도가", "수량",
                    "수익률(%)", "결과", "매도사유", "매수점수"]
        widths = [10, 12, 12, 12, 12, 8, 12, 8, 16, 10]

        for col, (h, w) in enumerate(zip(headers, widths), 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border
            ws2.column_dimensions[get_column_letter(col)].width = w

        for i, t in enumerate(result.get("trades", []), start=2):
            vals = [
                t["code"], t["buy_date"], t["sell_date"],
                t["buy_price"], t["sell_price"], t["qty"],
                t["pnl"], t["result"], t.get("reason", ""),
                t.get("score", 0),
            ]
            row_fill = win_fill if t["pnl"] > 0 else loss_fill
            for col, v in enumerate(vals, 1):
                cell = ws2.cell(row=i, column=col, value=v)
                cell.font = data_font
                cell.fill = row_fill
                cell.border = border
                if col in (4, 5):
                    cell.number_format = "#,##0"
                elif col == 7:
                    cell.number_format = "+0.00;-0.00;0"
                cell.alignment = center

        # ── Sheet 3: 일별 자산 ──
        ws3 = wb.create_sheet("일별 자산")
        eq_headers = ["날짜", "현금", "총자산", "수익률(%)", "보유종목"]
        for col, h in enumerate(eq_headers, 1):
            cell = ws3.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

        ws3.column_dimensions["A"].width = 12
        ws3.column_dimensions["B"].width = 15
        ws3.column_dimensions["C"].width = 15
        ws3.column_dimensions["D"].width = 12
        ws3.column_dimensions["E"].width = 10

        for i, eq in enumerate(result.get("daily_equity", []), start=2):
            ret = (eq["equity"] - self.initial_cash) / self.initial_cash * 100
            vals = [eq["date"], eq["cash"], eq["equity"], round(ret, 2), eq["holdings"]]
            for col, v in enumerate(vals, 1):
                cell = ws3.cell(row=i, column=col, value=v)
                cell.font = data_font
                cell.border = border
                if col in (2, 3):
                    cell.number_format = "#,##0"
                elif col == 4:
                    cell.number_format = "+0.00;-0.00;0"
                cell.alignment = center

        report_dir = _get_report_path()
        filename = f"backtest_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(report_dir, filename)
        wb.save(filepath)
        self._msg(f"엑셀 저장: {filepath}")
        return filepath
