"""
매매 성과 엑셀 리포트 자동 생성
- 일별 매매 내역
- 주별/월별 요약 통계
- 승률, 평균 수익률, 최대 손실 등
"""
import os
import sys
from datetime import datetime, timedelta
from ..db.database import get_connection

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def _get_report_path():
    if getattr(sys, 'frozen', False):
        base = os.path.dirname(os.path.dirname(sys.executable))
    else:
        base = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    report_dir = os.path.join(base, "reports")
    os.makedirs(report_dir, exist_ok=True)
    return report_dir


def _fetch_trades(days=None):
    conn = get_connection()
    try:
        if days:
            since = (datetime.now() - timedelta(days=days)).strftime("%Y%m%d")
            rows = conn.execute("""
                SELECT * FROM trade_results
                WHERE sell_date IS NOT NULL AND sell_date >= ?
                ORDER BY sell_date DESC, buy_date DESC
            """, (since,)).fetchall()
        else:
            rows = conn.execute("""
                SELECT * FROM trade_results
                WHERE sell_date IS NOT NULL
                ORDER BY sell_date DESC, buy_date DESC
            """).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


def _fetch_open_positions():
    conn = get_connection()
    try:
        rows = conn.execute("""
            SELECT * FROM trade_results
            WHERE sell_date IS NULL
            ORDER BY buy_date DESC
        """).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


def _calc_stats(trades):
    if not trades:
        return {
            "total": 0, "wins": 0, "losses": 0, "win_rate": 0,
            "avg_pnl": 0, "max_profit": 0, "max_loss": 0,
            "total_pnl": 0, "avg_hold_days": 0
        }

    wins = [t for t in trades if t.get("pnl", 0) > 0]
    losses = [t for t in trades if t.get("pnl", 0) <= 0]
    pnls = [t.get("pnl", 0) for t in trades]

    hold_days = []
    for t in trades:
        try:
            bd = datetime.strptime(t["buy_date"], "%Y%m%d")
            sd = datetime.strptime(t["sell_date"], "%Y%m%d")
            hold_days.append((sd - bd).days)
        except Exception:
            pass

    return {
        "total": len(trades),
        "wins": len(wins),
        "losses": len(losses),
        "win_rate": len(wins) / len(trades) * 100 if trades else 0,
        "avg_pnl": sum(pnls) / len(pnls) if pnls else 0,
        "max_profit": max(pnls) if pnls else 0,
        "max_loss": min(pnls) if pnls else 0,
        "total_pnl": sum(pnls),
        "avg_hold_days": sum(hold_days) / len(hold_days) if hold_days else 0,
    }


def generate_report(days=None):
    if not HAS_OPENPYXL:
        print("[리포트] openpyxl 미설치 — pip install openpyxl")
        return None

    trades = _fetch_trades(days)
    open_pos = _fetch_open_positions()
    stats = _calc_stats(trades)

    wb = Workbook()

    # ── 스타일 정의 ──
    header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2D3436")
    data_font = Font(name="Arial", size=10)
    win_fill = PatternFill("solid", fgColor="DFF9FB")
    loss_fill = PatternFill("solid", fgColor="FFE0E0")
    border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    center = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    # ════════════════════════════════════
    #  Sheet 1: 요약
    # ════════════════════════════════════
    ws = wb.active
    ws.title = "요약"

    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = f"AI 매매 성과 리포트 ({datetime.now().strftime('%Y-%m-%d')})"
    title_cell.font = Font(name="Arial", bold=True, size=14, color="2D3436")

    summary_items = [
        ("총 매매 횟수", stats["total"], "회"),
        ("승리", stats["wins"], "회"),
        ("패배", stats["losses"], "회"),
        ("승률", f"{stats['win_rate']:.1f}", "%"),
        ("평균 수익률", f"{stats['avg_pnl']:+.2f}", "%"),
        ("최대 수익", f"{stats['max_profit']:+.2f}", "%"),
        ("최대 손실", f"{stats['max_loss']:+.2f}", "%"),
        ("누적 수익률 합계", f"{stats['total_pnl']:+.2f}", "%"),
        ("평균 보유일수", f"{stats['avg_hold_days']:.1f}", "일"),
        ("미청산 포지션", len(open_pos), "건"),
    ]

    for i, (label, value, unit) in enumerate(summary_items, start=3):
        ws[f"A{i}"] = label
        ws[f"A{i}"].font = Font(name="Arial", bold=True, size=10)
        ws[f"B{i}"] = value
        ws[f"B{i}"].font = data_font
        ws[f"B{i}"].alignment = right_align
        ws[f"C{i}"] = unit
        ws[f"C{i}"].font = data_font

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 5

    # ── 주별 통계 ──
    if trades:
        ws.merge_cells("A15:F15")
        ws["A15"].value = "주별 통계"
        ws["A15"].font = Font(name="Arial", bold=True, size=12)

        weekly = {}
        for t in trades:
            try:
                sd = datetime.strptime(t["sell_date"], "%Y%m%d")
                week_key = f"{sd.isocalendar()[0]}-W{sd.isocalendar()[1]:02d}"
                if week_key not in weekly:
                    weekly[week_key] = []
                weekly[week_key].append(t)
            except Exception:
                pass

        w_headers = ["주차", "매매수", "승", "패", "승률(%)", "평균수익률(%)"]
        for col, h in enumerate(w_headers, 1):
            cell = ws.cell(row=16, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

        row = 17
        for week in sorted(weekly.keys(), reverse=True):
            wt = weekly[week]
            ws_stats = _calc_stats(wt)
            vals = [week, ws_stats["total"], ws_stats["wins"], ws_stats["losses"],
                    round(ws_stats["win_rate"], 1), round(ws_stats["avg_pnl"], 2)]
            for col, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=col, value=v)
                cell.font = data_font
                cell.alignment = center
                cell.border = border
            row += 1

    # ════════════════════════════════════
    #  Sheet 2: 매매 내역
    # ════════════════════════════════════
    ws2 = wb.create_sheet("매매 내역")
    headers = ["종목명", "코드", "매수일", "매도일", "매수가", "매도가",
               "수량", "수익률(%)", "결과", "AI점수", "보유일"]
    col_widths = [14, 10, 12, 12, 12, 12, 8, 12, 8, 10, 8]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws2.column_dimensions[get_column_letter(col)].width = w

    for i, t in enumerate(trades, start=2):
        hold_d = ""
        try:
            bd = datetime.strptime(t["buy_date"], "%Y%m%d")
            sd = datetime.strptime(t["sell_date"], "%Y%m%d")
            hold_d = (sd - bd).days
        except Exception:
            pass

        vals = [
            t.get("name", ""),
            t.get("code", ""),
            t.get("buy_date", ""),
            t.get("sell_date", ""),
            t.get("buy_price", 0),
            t.get("sell_price", 0),
            t.get("qty", 0),
            t.get("pnl", 0),
            t.get("result", ""),
            t.get("signal_score", 0),
            hold_d,
        ]
        row_fill = win_fill if t.get("pnl", 0) > 0 else loss_fill
        for col, v in enumerate(vals, 1):
            cell = ws2.cell(row=i, column=col, value=v)
            cell.font = data_font
            cell.fill = row_fill
            cell.border = border
            if col in (5, 6):
                cell.number_format = "#,##0"
            elif col == 8:
                cell.number_format = "+0.00;-0.00;0"
            cell.alignment = center if col not in (1,) else Alignment(vertical="center")

    # ════════════════════════════════════
    #  Sheet 3: 미청산 포지션
    # ════════════════════════════════════
    if open_pos:
        ws3 = wb.create_sheet("미청산")
        o_headers = ["종목명", "코드", "매수일", "매수가", "수량", "AI점수", "보유일"]
        o_widths = [14, 10, 12, 12, 8, 10, 8]

        for col, (h, w) in enumerate(zip(o_headers, o_widths), 1):
            cell = ws3.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border
            ws3.column_dimensions[get_column_letter(col)].width = w

        for i, t in enumerate(open_pos, start=2):
            hold_d = ""
            try:
                bd = datetime.strptime(t["buy_date"], "%Y%m%d")
                hold_d = (datetime.now() - bd).days
            except Exception:
                pass

            vals = [
                t.get("name", ""),
                t.get("code", ""),
                t.get("buy_date", ""),
                t.get("buy_price", 0),
                t.get("qty", 0),
                t.get("signal_score", 0),
                hold_d,
            ]
            for col, v in enumerate(vals, 1):
                cell = ws3.cell(row=i, column=col, value=v)
                cell.font = data_font
                cell.border = border
                if col == 4:
                    cell.number_format = "#,##0"
                cell.alignment = center if col != 1 else Alignment(vertical="center")

    # ── 저장 ──
    report_dir = _get_report_path()
    filename = f"trade_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(report_dir, filename)
    wb.save(filepath)
    print(f"[리포트] 엑셀 리포트 생성: {filepath}")
    return filepath
